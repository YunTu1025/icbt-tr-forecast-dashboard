import streamlit as st
import sys
import os
from pathlib import Path
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import datetime
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import SeriesLabel

os.environ['JAVA_HOME'] = r'C:\Program Files\Java\jdk-22'

# Increase pandas Styler render limit for large datasets
pd.set_option("styler.render.max_elements", 10000000)

sys.path.append(str(Path(__file__).parent))

from modules.data_loader import load_all_data
from modules.hive_connector import HiveConnection
from modules.hive_queries import HiveQueries
from modules.ml_forecaster import forecast_variable_tr_2026, forecast_international_tr_2026
from modules.fee_forecast_ui import render_fee_forecast_section, render_fee_forecast_section_run_rate
from modules.fee_forecast_engine import (
    FEE_TYPE_CONFIG,
    load_and_aggregate_data,
    build_forecast_table,
    calculate_baseline_row,
    build_index_table,
    update_forecast_2026_baseline
)
from modules.if_ff_store_ui import render_tab3, prepare_tab3_data_for_session
from modules.forecast_review_ui import render_fvf_daily_section, render_fvf_monthly_section, render_fvf_weekly_table_tab2
from modules.forecast_review_engine import build_fvf_weekly_table
from modules.tr_walk_section import display_walk_section

st.set_page_config(
    page_title="iCBT Take Rate Forecast Dashboard",
    page_icon="📊",
    layout="wide"
)

def main():
    st.title("iCBT Take Rate Forecast Dashboard")
    st.markdown("---")

    # Apply global CSS for right and middle alignment
    st.markdown("""
        <style>
            .stDataFrame th, .stDataFrame td {
                text-align: right !important;
                vertical-align: middle !important;
            }
        </style>
    """, unsafe_allow_html=True)

    # Initialize session state
    if 'hive_conn' not in st.session_state:
        st.session_state.hive_conn = None
        st.session_state.hive_connected = False
        st.session_state.fvf_data = None
        st.session_state.if_ff_store_data = None
        st.session_state.all_rev_data = None
        st.session_state.data_loaded = False
        st.session_state.data_source = 'Hive Database'  # Default data source

    # Initialize tab2_forecast_data separately (always check)
    if 'tab2_forecast_data' not in st.session_state:
        st.session_state.tab2_forecast_data = {}  # Store Tab2 forecast results for Tab4

    # Data Source Selection
    st.sidebar.header("📂 Data Source")
    data_source = st.sidebar.radio(
        "Choose data source:",
        ['Hive Database', 'CSV File Upload'],
        key='data_source_selector'
    )

    st.sidebar.markdown("---")

    # ==================== HIVE DATABASE CONNECTION ====================
    if data_source == 'Hive Database':
        st.sidebar.header("🔐 Hive Connection")
        st.sidebar.markdown("Enter your Hive credentials")

        username = st.sidebar.text_input("NT Username", key="hive_username")
        password = st.sidebar.text_input("PET Password", type="password", key="hive_password")

        if st.sidebar.button("📊 Retrieve Data", type="primary"):
            if not username or not password:
                st.sidebar.error("❌ Please enter NT Username and PET Password")
            else:
                with st.spinner("Connecting to Hive and retrieving data..."):
                    try:
                        hive_conn = HiveConnection(username, password)
                        success, message = hive_conn.connect()

                        if not success:
                            st.session_state.hive_connected = False
                            st.sidebar.error(f"❌ {message}")
                        else:
                            st.session_state.hive_conn = hive_conn
                            st.session_state.hive_connected = True

                            fvf_success = False
                            if_ff_success = False
                            all_rev_success = False
                            error_tables = []

                            try:
                                with st.spinner("Retrieving FVF data from P_FPA_T.BUDGET_DATA_OUTPUT..."):
                                    fvf_results = hive_conn.execute_query(HiveQueries.get_fvf_data())
                                    fvf_columns = hive_conn.get_column_names()
                                    st.session_state.fvf_data = pd.DataFrame(fvf_results, columns=fvf_columns)
                                    fvf_success = True
                                    st.sidebar.success(f"✅ FVF Data: {len(fvf_results):,} rows")
                            except Exception as e:
                                error_msg = str(e)
                                if "access" in error_msg.lower() or "permission" in error_msg.lower() or "denied" in error_msg.lower():
                                    error_tables.append("P_FPA_T.BUDGET_DATA_OUTPUT")
                                    st.sidebar.error(f"❌ Access denied to P_FPA_T.BUDGET_DATA_OUTPUT")
                                else:
                                    st.sidebar.error(f"❌ FVF query error: {error_msg[:100]}")

                            try:
                                with st.spinner("Retrieving IF/FF/Store data from P_FPA_T.FEE_FORECAST2..."):
                                    if_ff_results = hive_conn.execute_query(HiveQueries.get_if_ff_store_data())
                                    if_ff_columns = hive_conn.get_column_names()
                                    st.session_state.if_ff_store_data = pd.DataFrame(if_ff_results, columns=if_ff_columns)
                                    if_ff_success = True
                                    # Reset Tab3 data preparation flag when new data is loaded
                                    st.session_state['tab3_data_prepared'] = False
                                    st.sidebar.success(f"✅ IF/FF/Store Data: {len(if_ff_results):,} rows")
                            except Exception as e:
                                error_msg = str(e)
                                if "access" in error_msg.lower() or "permission" in error_msg.lower() or "denied" in error_msg.lower():
                                    error_tables.append("P_FPA_T.FEE_FORECAST2")
                                    st.sidebar.error(f"❌ Access denied to P_FPA_T.FEE_FORECAST2")
                                else:
                                    st.sidebar.error(f"❌ IF/FF query error: {error_msg[:100]}")

                            try:
                                with st.spinner("Retrieving ALL REV data from P_FPA_T.FEE_FORECAST2..."):
                                    all_rev_results = hive_conn.execute_query(HiveQueries.get_all_rev_data())
                                    all_rev_columns = hive_conn.get_column_names()
                                    st.session_state.all_rev_data = pd.DataFrame(all_rev_results, columns=all_rev_columns)
                                    all_rev_success = True
                                    st.sidebar.success(f"✅ ALL REV Data: {len(all_rev_results):,} rows")
                            except Exception as e:
                                error_msg = str(e)
                                if "access" in error_msg.lower() or "permission" in error_msg.lower() or "denied" in error_msg.lower():
                                    error_tables.append("P_FPA_T.FEE_FORECAST2 (ALL REV)")
                                    st.sidebar.error(f"❌ Access denied to P_FPA_T.FEE_FORECAST2 (ALL REV)")
                                else:
                                    st.sidebar.error(f"❌ ALL REV query error: {error_msg[:100]}")

                            if error_tables:
                                st.sidebar.warning(f"⚠️ Data access needed for: {', '.join(error_tables)}")

                            if fvf_success or if_ff_success or all_rev_success:
                                st.session_state.data_loaded = True
                                st.session_state.data_source = 'Hive Database'

                    except Exception as e:
                        st.session_state.hive_connected = False
                        st.sidebar.error(f"❌ Connection Error: {str(e)}")

        if st.session_state.data_loaded and st.session_state.data_source == 'Hive Database':
            st.sidebar.success("✅ Data Retrieved")
        elif st.session_state.hive_connected:
            st.sidebar.info("ℹ️ Connected - Click 'Retrieve Data' to load")
        else:
            st.sidebar.warning("⚠️ Not Connected")

    # ==================== CSV FILE UPLOAD ====================
    else:  # CSV File Upload
        st.sidebar.header("📤 CSV File Upload")
        st.sidebar.markdown("Upload CSV files for FVF, IF/FF/Store, and ALL REV data")

        fvf_csv_file = st.sidebar.file_uploader(
            "Upload FVF Data (CSV)",
            type=['csv'],
            accept_multiple_files=False,
            key='fvf_csv_uploader'
        )

        if_ff_csv_file = st.sidebar.file_uploader(
            "Upload IF/FF/Store Data (CSV)",
            type=['csv'],
            accept_multiple_files=False,
            key='if_ff_csv_uploader'
        )

        all_rev_csv_file = st.sidebar.file_uploader(
            "Upload ALL REV Data (CSV)",
            type=['csv'],
            accept_multiple_files=False,
            key='all_rev_csv_uploader'
        )

        if st.sidebar.button("📊 Load CSV Data", type="primary"):
            if not fvf_csv_file and not if_ff_csv_file and not all_rev_csv_file:
                st.sidebar.error("❌ Please upload at least one CSV file")
            else:
                with st.spinner("Loading CSV files..."):
                    try:
                        fvf_success = False
                        if_ff_success = False
                        all_rev_success = False

                        # Load FVF CSV
                        if fvf_csv_file:
                            try:
                                st.session_state.fvf_data = pd.read_csv(fvf_csv_file)
                                fvf_success = True
                                st.sidebar.success(f"✅ FVF Data: {len(st.session_state.fvf_data):,} rows")
                            except Exception as e:
                                st.sidebar.error(f"❌ FVF CSV error: {str(e)[:100]}")

                        # Load IF/FF/Store CSV
                        if if_ff_csv_file:
                            try:
                                st.session_state.if_ff_store_data = pd.read_csv(if_ff_csv_file)
                                if_ff_success = True
                                # Reset Tab3 data preparation flag when new data is loaded
                                st.session_state['tab3_data_prepared'] = False
                                st.sidebar.success(f"✅ IF/FF/Store Data: {len(st.session_state.if_ff_store_data):,} rows")
                            except Exception as e:
                                st.sidebar.error(f"❌ IF/FF CSV error: {str(e)[:100]}")

                        # Load ALL REV CSV
                        if all_rev_csv_file:
                            try:
                                st.session_state.all_rev_data = pd.read_csv(all_rev_csv_file)
                                all_rev_success = True
                                st.sidebar.success(f"✅ ALL REV Data: {len(st.session_state.all_rev_data):,} rows")
                            except Exception as e:
                                st.sidebar.error(f"❌ ALL REV CSV error: {str(e)[:100]}")

                        if fvf_success or if_ff_success or all_rev_success:
                            st.session_state.data_loaded = True
                            st.session_state.data_source = 'CSV File Upload'

                    except Exception as e:
                        st.sidebar.error(f"❌ Error loading CSV: {str(e)}")

        if st.session_state.data_loaded and st.session_state.data_source == 'CSV File Upload':
            st.sidebar.success("✅ CSV Data Loaded")

    if not st.session_state.data_loaded:
        st.info("👈 Select a data source in the sidebar (Hive Database or CSV File Upload) to load data")

        st.markdown("### 📋 Data Source Options")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### 🗄️ Hive Database")
            st.markdown("""
            **Available Data Sources:**

            1. **FVF Data**
               - Table: `P_FPA_T.BUDGET_DATA_OUTPUT`
               - Filter: `RETAIL_YEAR >= 2021`
               - Used for: Final Value Fee forecast calculations

            2. **IF/FF/Store Data**
               - Table: `P_FPA_T.FEE_FORECAST2`
               - Filter: `DT >= '2021-01-01'`
               - Aggregated by date, year, month, week, and country
               - Used for: Insertion Fee, Feature Fee, Store Fee analysis

            3. **ALL REV Data**
               - Table: `P_FPA_T.FEE_FORECAST2`
               - Filter: `YEAR_ID = 2026`
               - Used for: Comprehensive revenue analysis

            **Steps:**
            1. Enter your NT Username
            2. Enter your PET Password
            3. Click "📊 Retrieve Data"
            """)

        with col2:
            st.markdown("#### 📁 CSV File Upload")
            st.markdown("""
            **Upload Options:**

            You can upload one or more of the following:

            1. **FVF Data (CSV)** - Optional
               - Final Value Fee forecast data
               - Should match `P_FPA_T.BUDGET_DATA_OUTPUT` schema

            2. **IF/FF/Store Data (CSV)** - Optional
               - Insertion/Feature/Store fee data
               - Should match aggregated `P_FPA_T.FEE_FORECAST2` schema

            3. **ALL REV Data (CSV)** - Optional
               - Complete revenue data
               - Should match `P_FPA_T.FEE_FORECAST2` schema

            **Steps:**
            1. Click "Upload" for each data source you want to upload
            2. Select your CSV file. Only one CSV file is allowed
            3. Click "📊 Load CSV Data"

            **Note:**
            - CSV files should have proper column headers
            """)

        return

    def format_dataframe_for_display(df, metric_cols_start, metric_cols_end):
        df_display = df.copy()

        # Reset index to start from 1
        df_display = df_display.reset_index(drop=True)
        df_display.index = df_display.index + 1

        # Format numeric columns with thousand separators
        for col_idx in range(metric_cols_start, min(metric_cols_end + 1, len(df.columns))):
            col_name = df.columns[col_idx]
            if pd.api.types.is_numeric_dtype(df_display[col_name]):
                df_display[col_name] = df_display[col_name].apply(
                    lambda x: f"{x:,.0f}" if pd.notna(x) else ""
                )

        return df_display

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📊 Data Overview",
        "💰 FVF Forecast",
        "📈 IF/FF/Store Forecast",
        "📋 Forecast Result Review",
        "🔄 TR Walk",
        "📌 PLUG and Submission"
    ])

    with tab1:
        if st.session_state.fvf_data is not None:
            st.markdown("### FVF Data Sample")
            fvf_sorted = st.session_state.fvf_data.sort_values(
                by=[st.session_state.fvf_data.columns[0], st.session_state.fvf_data.columns[1]],
                ascending=[False, False]
            )
            fvf_display = format_dataframe_for_display(fvf_sorted.head(100), 9, 48)
            st.dataframe(fvf_display, use_container_width=True, height=500)
            st.caption(f"Showing {min(len(fvf_sorted), 100)} of {len(fvf_sorted):,} rows")
        else:
            st.warning("⚠️ FVF data not available - access denied to P_FPA_T.BUDGET_DATA_OUTPUT")

        st.markdown("---")

        if st.session_state.if_ff_store_data is not None:
            st.markdown("### IF/FF/Store Data Sample")
            if_ff_sorted = st.session_state.if_ff_store_data.sort_values(
                by=st.session_state.if_ff_store_data.columns[0],
                ascending=False
            )
            if_ff_display = format_dataframe_for_display(if_ff_sorted.head(100), 9, 13)
            st.dataframe(if_ff_display, use_container_width=True, height=500)
            st.caption(f"Showing {min(len(if_ff_sorted), 100)} of {len(if_ff_sorted):,} rows")
        else:
            st.warning("⚠️ IF/FF/Store data not available - access denied to P_FPA_T.FEE_FORECAST2")

        st.markdown("---")

        if st.session_state.all_rev_data is not None:
            st.markdown("### ALL REV Data Sample")
            all_rev_sorted = st.session_state.all_rev_data.sort_values(
                by='DT',
                ascending=False
            )
            # Format all metric columns from column 9 to the last column
            all_rev_display = format_dataframe_for_display(all_rev_sorted.head(100), 9, len(all_rev_sorted.columns) - 1)
            st.dataframe(all_rev_display, use_container_width=True, height=500)
            st.caption(f"Showing {min(len(all_rev_sorted), 100)} of {len(all_rev_sorted):,} rows")
        else:
            st.warning("⚠️ ALL REV data not available - access denied to P_FPA_T.FEE_FORECAST2")

    with tab2:
        if st.session_state.fvf_data is None:
            st.error("❌ No access to P_FPA_T.BUDGET_DATA_OUTPUT")
            st.info("Please contact your administrator to request access to this table.")
            return

        # Define fee types in specified order
        fee_types_list = [
            "Variable",
            "International",
            "BSTD",
            "eTRS",
            "SNAD",
            "Fixed",
            "Credit",
            "Regulatory",
            "Buyer Protection"
        ]

        # Years available for weighting
        years_list = [2022, 2023, 2024, 2025]

        # Initialize control panel settings in session state
        if 'control_panel_settings' not in st.session_state:
            st.session_state.control_panel_settings = {
                'current_year': 2026,
                'index_weeks': [],
                'fee_type_weights': {}
            }

        # Initialize fee type weights with new names if not exists or if old names exist
        if not st.session_state.control_panel_settings['fee_type_weights'] or \
           any(key not in fee_types_list for key in st.session_state.control_panel_settings['fee_type_weights'].keys()):
            st.session_state.control_panel_settings['fee_type_weights'] = {}
            for fee_type in fee_types_list:
                st.session_state.control_panel_settings['fee_type_weights'][fee_type] = {
                    'weights': {year: 100.0/len(years_list) for year in years_list}
                }

        # Section 1: Control Panel
        st.markdown("### 📋 Control Panel")

        # Hardcoded week to date mapping (from Control Panel A15:B66)
        week_date_map = {
            1: '2026-01-04', 2: '2026-01-11', 3: '2026-01-18', 4: '2026-01-25',
            5: '2026-02-01', 6: '2026-02-08', 7: '2026-02-15', 8: '2026-02-22',
            9: '2026-03-01', 10: '2026-03-08', 11: '2026-03-15', 12: '2026-03-22',
            13: '2026-03-29', 14: '2026-04-05', 15: '2026-04-12', 16: '2026-04-19',
            17: '2026-04-26', 18: '2026-05-03', 19: '2026-05-10', 20: '2026-05-17',
            21: '2026-05-24', 22: '2026-05-31', 23: '2026-06-07', 24: '2026-06-14',
            25: '2026-06-21', 26: '2026-06-28', 27: '2026-07-05', 28: '2026-07-12',
            29: '2026-07-19', 30: '2026-07-26', 31: '2026-08-02', 32: '2026-08-09',
            33: '2026-08-16', 34: '2026-08-23', 35: '2026-08-30', 36: '2026-09-06',
            37: '2026-09-13', 38: '2026-09-20', 39: '2026-09-27', 40: '2026-10-04',
            41: '2026-10-11', 42: '2026-10-18', 43: '2026-10-25', 44: '2026-11-01',
            45: '2026-11-08', 46: '2026-11-15', 47: '2026-11-22', 48: '2026-11-29',
            49: '2026-12-06', 50: '2026-12-13', 51: '2026-12-20', 52: '2026-12-27'
        }

        available_weeks = sorted(week_date_map.keys())

        # All inputs in one row
        col1a, col1b, col2a, col2b, col3a, col3b, col4 = st.columns([0.5, 0.8, 0.5, 0.8, 0.5, 0.8, 2.1])

        with col1a:
            st.markdown("<div style='margin-top: 8px;'>Current Year</div>", unsafe_allow_html=True)
        with col1b:
            current_year = st.number_input(
                "Current Year",
                min_value=2020,
                max_value=2030,
                value=st.session_state.control_panel_settings['current_year'],
                step=1,
                key="current_year_input",
                label_visibility="collapsed"
            )
            st.session_state.control_panel_settings['current_year'] = current_year

        # Calculate default weeks based on max week of current year in FVF data
        if st.session_state.fvf_data is not None and 'RETAIL_WEEK' in st.session_state.fvf_data.columns and 'RETAIL_YEAR' in st.session_state.fvf_data.columns:
            # Filter by current year, then get max week
            current_year_data = st.session_state.fvf_data[st.session_state.fvf_data['RETAIL_YEAR'] == current_year]
            if len(current_year_data) > 0:
                current_week = int(current_year_data['RETAIL_WEEK'].max())
            else:
                current_week = max(available_weeks) if available_weeks else 1  # Fallback if no data for current year
        else:
            current_week = max(available_weeks) if available_weeks else 1  # Fallback to hardcoded weeks

        default_start_week = max(1, current_week - 3)  # current week - 3, but not less than 1
        default_end_week = current_week

        with col2a:
            st.markdown("<div style='margin-top: 8px;'>Start Week</div>", unsafe_allow_html=True)
        with col2b:
            week_start = st.selectbox(
                "Start Week",
                options=available_weeks,
                index=available_weeks.index(default_start_week) if default_start_week in available_weeks else 0,
                key="index_week_start",
                label_visibility="collapsed"
            )

        with col3a:
            st.markdown("<div style='margin-top: 8px;'>End Week</div>", unsafe_allow_html=True)
        with col3b:
            available_end_weeks = [w for w in available_weeks if w >= week_start]
            week_end = st.selectbox(
                "End Week",
                options=available_end_weeks,
                index=available_end_weeks.index(default_end_week) if default_end_week in available_end_weeks else (len(available_end_weeks)-1 if len(available_end_weeks) > 0 else 0),
                key="index_week_end",
                label_visibility="collapsed"
            )

        with col4:
            st.markdown(f"<div style='margin-top: 8px;'><strong>Week Begin Dates:</strong> Start: <strong><u>{week_date_map[week_start]}</u></strong> ; End: <strong><u>{week_date_map[week_end]}</u></strong></div>", unsafe_allow_html=True)

        st.session_state.control_panel_settings['index_weeks'] = list(range(week_start, week_end + 1))
        st.info(f"Selected {len(st.session_state.control_panel_settings['index_weeks'])} weeks for index calculation")

        # ===== PREVIOUS FORECAST VERSION SECTION =====
        st.markdown("---")
        st.markdown("### 📂 Previous Forecast Version")

        # Upload areas in a single row with 4 columns
        upload_col1, upload_col2, upload_col3, upload_col4 = st.columns([0.5, 1.5, 0.7, 1.5])

        with upload_col1:
            st.markdown("<div style='margin-top: 16px;'>Budget</div>", unsafe_allow_html=True)

        with upload_col2:
            budget_file = st.file_uploader(
                "Upload Budget CSV",
                type=['csv'],
                accept_multiple_files=False,
                key="budget_csv_upload",
                label_visibility="collapsed"
            )

        with upload_col3:
            st.markdown("<div style='margin-top: 16px;'>Prior Forecast</div>", unsafe_allow_html=True)

        with upload_col4:
            prior_forecast_file = st.file_uploader(
                "Upload Prior Forecast CSV",
                type=['csv'],
                accept_multiple_files=False,
                key="prior_forecast_csv_upload",
                label_visibility="collapsed"
            )

        # Success messages in second row
        success_col1, success_col2 = st.columns(2)

        with success_col1:
            if budget_file is not None:
                st.success(f"✅ {budget_file.name} uploaded")

        with success_col2:
            if prior_forecast_file is not None:
                st.success(f"✅ {prior_forecast_file.name} uploaded")

        # Read uploaded CSV files
        budget_df = None
        prior_forecast_df = None

        # Display uploaded file contents in tables
        table_col1, table_col2 = st.columns(2)

        with table_col1:
            if budget_file is not None:
                try:
                    budget_df = pd.read_csv(budget_file)
                    # Format all columns except Week/RETAIL_WEEK as percentage
                    budget_df_display = budget_df.copy()
                    budget_df_display = budget_df_display.reset_index(drop=True)
                    for col in budget_df_display.columns:
                        if col not in ['Week', 'RETAIL_WEEK']:
                            budget_df_display[col] = budget_df_display[col].apply(
                                lambda x: f"{x*100:.2f}%" if pd.notna(x) and isinstance(x, (int, float)) else x
                            )
                    st.markdown("##### Budget Data Preview")
                    st.dataframe(budget_df_display, use_container_width=True, height=300, hide_index=True)
                    st.caption(f"Total Records: {len(budget_df):,}")
                except Exception as e:
                    st.error(f"Error reading Budget CSV: {str(e)}")
                    budget_df = None

        with table_col2:
            if prior_forecast_file is not None:
                try:
                    prior_forecast_df = pd.read_csv(prior_forecast_file)
                    # Format all columns except Week/RETAIL_WEEK as percentage
                    prior_forecast_df_display = prior_forecast_df.copy()
                    prior_forecast_df_display = prior_forecast_df_display.reset_index(drop=True)
                    for col in prior_forecast_df_display.columns:
                        if col not in ['Week', 'RETAIL_WEEK']:
                            prior_forecast_df_display[col] = prior_forecast_df_display[col].apply(
                                lambda x: f"{x*100:.2f}%" if pd.notna(x) and isinstance(x, (int, float)) else x
                            )
                    st.markdown("##### Prior Forecast Data Preview")
                    st.dataframe(prior_forecast_df_display, use_container_width=True, height=300, hide_index=True)
                    st.caption(f"Total Records: {len(prior_forecast_df):,}")
                except Exception as e:
                    st.error(f"Error reading Prior Forecast CSV: {str(e)}")
                    prior_forecast_df = None

        # ===== VARIABLE FORECAST SECTION =====
        if st.session_state.fvf_data is not None:
            df_variable_forecast = render_fee_forecast_section(
                fee_type='Variable',
                fvf_data=st.session_state.fvf_data,
                control_panel_settings=st.session_state.control_panel_settings,
                week_date_map=week_date_map,
                budget_df=budget_df,
                prior_forecast_df=prior_forecast_df
            )
            st.session_state.tab2_forecast_data['Variable'] = df_variable_forecast

        # ===== INTERNATIONAL FORECAST SECTION =====
        if st.session_state.fvf_data is not None:
            df_international_forecast = render_fee_forecast_section(
                fee_type='International',
                fvf_data=st.session_state.fvf_data,
                control_panel_settings=st.session_state.control_panel_settings,
                week_date_map=week_date_map,
                budget_df=budget_df,
                prior_forecast_df=prior_forecast_df
            )
            st.session_state.tab2_forecast_data['International'] = df_international_forecast

        # ===== BSTD FORECAST SECTION =====
        if st.session_state.fvf_data is not None:
            df_bstd_forecast = render_fee_forecast_section(
                fee_type='BSTD',
                fvf_data=st.session_state.fvf_data,
                control_panel_settings=st.session_state.control_panel_settings,
                week_date_map=week_date_map,
                budget_df=budget_df,
                prior_forecast_df=prior_forecast_df
            )
            st.session_state.tab2_forecast_data['BSTD'] = df_bstd_forecast

        # ===== eTRS FORECAST SECTION =====
        if st.session_state.fvf_data is not None:
            df_etrs_forecast = render_fee_forecast_section_run_rate(
                fee_type='eTRS',
                fvf_data=st.session_state.fvf_data,
                control_panel_settings=st.session_state.control_panel_settings,
                week_date_map=week_date_map,
                budget_df=budget_df,
                prior_forecast_df=prior_forecast_df
            )
            st.session_state.tab2_forecast_data['eTRS'] = df_etrs_forecast

        # ===== SNAD FORECAST SECTION =====
        if st.session_state.fvf_data is not None:
            df_snad_forecast = render_fee_forecast_section(
                fee_type='SNAD',
                fvf_data=st.session_state.fvf_data,
                control_panel_settings=st.session_state.control_panel_settings,
                week_date_map=week_date_map,
                budget_df=budget_df,
                prior_forecast_df=prior_forecast_df
            )
            st.session_state.tab2_forecast_data['SNAD'] = df_snad_forecast

        # ===== FIXED FORECAST SECTION =====
        if st.session_state.fvf_data is not None:
            df_fixed_forecast = render_fee_forecast_section(
                fee_type='Fixed',
                fvf_data=st.session_state.fvf_data,
                control_panel_settings=st.session_state.control_panel_settings,
                week_date_map=week_date_map,
                budget_df=budget_df,
                prior_forecast_df=prior_forecast_df
            )
            st.session_state.tab2_forecast_data['Fixed'] = df_fixed_forecast

        # ===== CREDIT FORECAST SECTION =====
        if st.session_state.fvf_data is not None:
            df_credit_forecast = render_fee_forecast_section(
                fee_type='Credit',
                fvf_data=st.session_state.fvf_data,
                control_panel_settings=st.session_state.control_panel_settings,
                week_date_map=week_date_map,
                budget_df=budget_df,
                prior_forecast_df=prior_forecast_df
            )
            st.session_state.tab2_forecast_data['Credit'] = df_credit_forecast

        # ===== REGULATORY FORECAST SECTION =====
        if st.session_state.fvf_data is not None:
            df_regulatory_forecast = render_fee_forecast_section_run_rate(
                fee_type='Regulatory',
                fvf_data=st.session_state.fvf_data,
                control_panel_settings=st.session_state.control_panel_settings,
                week_date_map=week_date_map,
                budget_df=budget_df,
                prior_forecast_df=prior_forecast_df
            )
            st.session_state.tab2_forecast_data['Regulatory'] = df_regulatory_forecast

        # ===== BUYER PROTECTION FORECAST SECTION =====
        if st.session_state.fvf_data is not None:
            df_buyer_protection_forecast = render_fee_forecast_section_run_rate(
                fee_type='Buyer Protection',
                fvf_data=st.session_state.fvf_data,
                control_panel_settings=st.session_state.control_panel_settings,
                week_date_map=week_date_map,
                budget_df=budget_df,
                prior_forecast_df=prior_forecast_df
            )
            st.session_state.tab2_forecast_data['Buyer Protection'] = df_buyer_protection_forecast

            # Build and store weekly table data for use in Tab4
            # This ensures the data is available as soon as forecasts are generated
            if st.session_state.tab2_forecast_data:
                fvf_weekly_df = build_fvf_weekly_table(
                    forecast_data_dict=st.session_state.tab2_forecast_data,
                    week_date_map=week_date_map,
                    gmv_data=st.session_state.get('fvf_weekly_table_data', None)
                )
                st.session_state['fvf_weekly_table_data'] = fvf_weekly_df

        # Render FVF Weekly Forecast Summary Table before the download button
        st.markdown("---")
        if st.session_state.tab2_forecast_data:
            render_fvf_weekly_table_tab2(st.session_state.tab2_forecast_data, week_date_map)
        else:
            st.info("💡 Generate forecasts above to see the FVF Weekly Forecast Summary Table")

        # Download forecast data button
        st.markdown("---")
        if st.button("📥 Download Forecast Process", type="primary"):
            if st.session_state.fvf_data is not None:
                with st.spinner("Generating Excel file..."):
                    # Create Excel file in memory
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:

                        # 1. Create Control Panel sheet
                        control_panel_data = {
                            'Setting': ['Current Year', 'Start Week', 'End Week', 'Start Date', 'End Date', 'Number of Weeks'],
                            'Value': [
                                st.session_state.control_panel_settings['current_year'],
                                st.session_state.control_panel_settings['index_weeks'][0],
                                st.session_state.control_panel_settings['index_weeks'][-1],
                                week_date_map[st.session_state.control_panel_settings['index_weeks'][0]],
                                week_date_map[st.session_state.control_panel_settings['index_weeks'][-1]],
                                len(st.session_state.control_panel_settings['index_weeks'])
                            ]
                        }
                        pd.DataFrame(control_panel_data).to_excel(writer, sheet_name='Control_Panel', index=False)

                        # Fee types using Index Baseline methodology
                        index_baseline_fee_types = ['Variable', 'International', 'BSTD', 'Fixed', 'Credit', 'SNAD']

                        for fee_type in index_baseline_fee_types:
                            if fee_type in FEE_TYPE_CONFIG:
                                config = FEE_TYPE_CONFIG[fee_type]

                                # Load and aggregate data
                                df_agg, error_msg = load_and_aggregate_data(
                                    st.session_state.fvf_data, config, fee_type
                                )

                                if df_agg is not None and error_msg is None:
                                    # Get weights for this fee type
                                    if fee_type in st.session_state.control_panel_settings['fee_type_weights']:
                                        weights = st.session_state.control_panel_settings['fee_type_weights'][fee_type]['weights']
                                    else:
                                        weights = {2022: 25.0, 2023: 25.0, 2024: 25.0, 2025: 25.0}

                                    # 2. Create Year Weights sheet
                                    weights_data = {
                                        'Year': [2022, 2023, 2024, 2025],
                                        'Weight (%)': [weights[2022], weights[2023], weights[2024], weights[2025]]
                                    }
                                    pd.DataFrame(weights_data).to_excel(writer, sheet_name=f"{fee_type}_Weights", index=False)

                                    # Build forecast table
                                    df_forecast = build_forecast_table(df_agg, config, week_date_map)
                                    baseline_row = calculate_baseline_row(df_forecast, st.session_state.control_panel_settings['index_weeks'])
                                    df_forecast = pd.concat([pd.DataFrame([baseline_row]), df_forecast], ignore_index=True)

                                    # Build index table
                                    df_index = build_index_table(df_forecast, weights, week_date_map)

                                    # Update 2026 baseline
                                    baseline_2026 = df_forecast.iloc[0]['2026']
                                    update_forecast_2026_baseline(df_forecast, df_index, baseline_2026)

                                    # Write sheets
                                    df_agg.to_excel(writer, sheet_name=f"{fee_type}_Raw", index=False)
                                    df_forecast.to_excel(writer, sheet_name=f"{fee_type}_Forecast", index=False)
                                    df_index.to_excel(writer, sheet_name=f"{fee_type}_Index", index=False)

                        # Fee types using Run Rate methodology
                        run_rate_fee_types = ['eTRS', 'Regulatory', 'Buyer Protection']

                        for fee_type in run_rate_fee_types:
                            if fee_type in FEE_TYPE_CONFIG:
                                config = FEE_TYPE_CONFIG[fee_type]

                                # Load and aggregate data
                                df_agg, error_msg = load_and_aggregate_data(
                                    st.session_state.fvf_data, config, fee_type
                                )

                                if df_agg is not None and error_msg is None:
                                    # Build forecast table
                                    df_forecast = build_forecast_table(df_agg, config, week_date_map)
                                    baseline_row = calculate_baseline_row(df_forecast, st.session_state.control_panel_settings['index_weeks'])
                                    df_forecast = pd.concat([pd.DataFrame([baseline_row]), df_forecast], ignore_index=True)

                                    # Update 2026 baseline (Run Rate: use baseline value for all missing weeks)
                                    baseline_2026 = df_forecast.iloc[0]['2026']
                                    for i in range(1, len(df_forecast)):
                                        value_2026 = df_forecast.iloc[i]['2026']
                                        if pd.notna(value_2026):
                                            df_forecast.at[i, '2026 Index Baseline'] = None
                                        else:
                                            if pd.notna(baseline_2026):
                                                df_forecast.at[i, '2026 Index Baseline'] = baseline_2026
                                            else:
                                                df_forecast.at[i, '2026 Index Baseline'] = None

                                    # Write sheets (no index table for Run Rate)
                                    df_agg.to_excel(writer, sheet_name=f"{fee_type}_Raw", index=False)
                                    df_forecast.to_excel(writer, sheet_name=f"{fee_type}_Forecast", index=False)

                        # Get the workbook to add charts
                        workbook = writer.book

                        # 3. Add charts to Index Baseline fee types
                        for fee_type in index_baseline_fee_types:
                            if fee_type in FEE_TYPE_CONFIG:
                                forecast_sheet_name = f"{fee_type}_Forecast"
                                index_sheet_name = f"{fee_type}_Index"

                                if forecast_sheet_name in workbook.sheetnames:
                                    ws_forecast = workbook[forecast_sheet_name]

                                    # Create Forecast Chart
                                    chart_forecast = LineChart()
                                    chart_forecast.title = f"{fee_type} Forecast"
                                    chart_forecast.style = 2
                                    chart_forecast.y_axis.title = f"{fee_type} TR (%)"
                                    chart_forecast.x_axis.title = "Week"
                                    chart_forecast.height = 15
                                    chart_forecast.width = 30

                                    # Data for years 2022-2026, 2026 Index Baseline, 2026 Machine Learning (columns C-I, skip row 1 baseline)
                                    years_cols = ['2022', '2023', '2024', '2025', '2026', '2026 Index Baseline', '2026 Machine Learning']
                                    for idx, year in enumerate(years_cols):
                                        values = Reference(ws_forecast, min_col=3+idx, min_row=2, max_row=53)  # Skip baseline row
                                        chart_forecast.add_data(values, titles_from_data=False)

                                    # Set series titles
                                    for idx, series in enumerate(chart_forecast.series):
                                        series.title = SeriesLabel(v=years_cols[idx])

                                    # Set categories (Week numbers)
                                    cats = Reference(ws_forecast, min_col=1, min_row=3, max_row=53)
                                    chart_forecast.set_categories(cats)

                                    ws_forecast.add_chart(chart_forecast, "K2")

                                if index_sheet_name in workbook.sheetnames:
                                    ws_index = workbook[index_sheet_name]

                                    # Create Index Chart
                                    chart_index = LineChart()
                                    chart_index.title = f"{fee_type} Index"
                                    chart_index.style = 2
                                    chart_index.y_axis.title = "Index"
                                    chart_index.x_axis.title = "Week"
                                    chart_index.height = 15
                                    chart_index.width = 30

                                    # Data for years 2022-2026, 2026 Index Baseline, 2026 Machine Learning (skip row 1 header and row 2 weights)
                                    for idx, year in enumerate(years_cols):
                                        values = Reference(ws_index, min_col=3+idx, min_row=3, max_row=54)  # Start from row 3 to skip weights
                                        chart_index.add_data(values, titles_from_data=False)

                                    # Set series titles
                                    for idx, series in enumerate(chart_index.series):
                                        series.title = SeriesLabel(v=years_cols[idx])

                                    # Set categories (Week numbers)
                                    cats = Reference(ws_index, min_col=1, min_row=3, max_row=54)  # Week numbers start from row 3
                                    chart_index.set_categories(cats)

                                    ws_index.add_chart(chart_index, "K2")

                        # 4. Add charts to Run Rate fee types
                        for fee_type in run_rate_fee_types:
                            if fee_type in FEE_TYPE_CONFIG:
                                forecast_sheet_name = f"{fee_type}_Forecast"

                                if forecast_sheet_name in workbook.sheetnames:
                                    ws_forecast = workbook[forecast_sheet_name]

                                    # Create Forecast Chart
                                    chart_forecast = LineChart()
                                    chart_forecast.title = f"{fee_type} Forecast"
                                    chart_forecast.style = 2
                                    chart_forecast.y_axis.title = f"{fee_type} TR (%)"
                                    chart_forecast.x_axis.title = "Week"
                                    chart_forecast.height = 15
                                    chart_forecast.width = 30

                                    # Data for years 2022-2026, 2026 Index Baseline, 2026 Machine Learning
                                    for idx, year in enumerate(years_cols):
                                        values = Reference(ws_forecast, min_col=3+idx, min_row=2, max_row=53)
                                        chart_forecast.add_data(values, titles_from_data=False)

                                    # Set series titles
                                    for idx, series in enumerate(chart_forecast.series):
                                        series.title = SeriesLabel(v=years_cols[idx])

                                    # Set categories (Week numbers)
                                    cats = Reference(ws_forecast, min_col=1, min_row=3, max_row=53)
                                    chart_forecast.set_categories(cats)

                                    ws_forecast.add_chart(chart_forecast, "K2")

                    # Prepare download
                    output.seek(0)
                    st.download_button(
                        label="📥 Download Excel File",
                        data=output,
                        file_name="FVF_Forecast_Process.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                st.success("✅ Excel file generated successfully!")
            else:
                st.error("❌ No FVF data available. Please load data first.")

    # Prepare Tab3 data before rendering tabs (ensures IF daily data is available to Tab4)
    prepare_tab3_data_for_session(st.session_state.if_ff_store_data)

    with tab3:
        tab3_monthly_data = render_tab3(st.session_state.if_ff_store_data)

    with tab4:
        # Check if forecast data is available
        if not st.session_state.tab2_forecast_data:
            st.warning("⚠️ No forecast data available. Please generate forecasts in Tab2 first.")
        else:
            # Render FVF Daily forecast section
            render_fvf_daily_section(
                forecast_data_dict=st.session_state.tab2_forecast_data,
                week_date_map=week_date_map,
                tab3_monthly_data=tab3_monthly_data,
                all_rev_data=st.session_state.all_rev_data
            )

            st.markdown("---")

            # Render FVF Monthly forecast section
            render_fvf_monthly_section(
                tab3_monthly_data=tab3_monthly_data
            )

    # ===== TAB5: TR Walk - Forecast Version Comparison =====
    with tab5:
        # ===== PREVIOUS FORECAST VERSION SECTION =====
        st.markdown("### 📂 Previous Forecast Version")

        # Upload areas in a single row with 4 columns
        upload_col1, upload_col2, upload_col3, upload_col4 = st.columns([0.5, 1.5, 0.7, 1.5])

        with upload_col1:
            st.markdown("<div style='margin-top: 16px;'>Budget</div>", unsafe_allow_html=True)

        with upload_col2:
            tab5_budget_file = st.file_uploader(
                "Upload Budget CSV",
                type=['csv'],
                accept_multiple_files=False,
                key="tab5_budget_csv_upload",
                label_visibility="collapsed"
            )

        with upload_col3:
            st.markdown("<div style='margin-top: 16px;'>Prior Forecast</div>", unsafe_allow_html=True)

        with upload_col4:
            tab5_prior_forecast_file = st.file_uploader(
                "Upload Prior Forecast CSV",
                type=['csv'],
                accept_multiple_files=False,
                key="tab5_prior_forecast_csv_upload",
                label_visibility="collapsed"
            )

        # Success messages in second row
        success_col1, success_col2 = st.columns(2)

        with success_col1:
            if tab5_budget_file is not None:
                st.success(f"✅ {tab5_budget_file.name} uploaded")

        with success_col2:
            if tab5_prior_forecast_file is not None:
                st.success(f"✅ {tab5_prior_forecast_file.name} uploaded")

        # Read uploaded CSV files
        tab5_budget_df = None
        tab5_prior_forecast_df = None

        # Display uploaded file contents in tables
        table_col1, table_col2 = st.columns(2)

        with table_col1:
            if tab5_budget_file is not None:
                try:
                    tab5_budget_df = pd.read_csv(tab5_budget_file)
                    # Format all columns except Date/Week/Month as number
                    tab5_budget_df_display = tab5_budget_df.copy()
                    tab5_budget_df_display = tab5_budget_df_display.reset_index(drop=True)
                    for col in tab5_budget_df_display.columns:
                        if col not in ['Date', 'Week', 'Month']:
                            tab5_budget_df_display[col] = tab5_budget_df_display[col].apply(
                                lambda x: f"{x:,.0f}" if pd.notna(x) and isinstance(x, (int, float)) else x
                            )
                    st.markdown("##### Budget Data Preview")
                    st.dataframe(tab5_budget_df_display, use_container_width=True, height=480, hide_index=True)
                    st.caption(f"Total Records: {len(tab5_budget_df):,}")
                except Exception as e:
                    st.error(f"Error reading Budget CSV: {str(e)}")
                    tab5_budget_df = None

        with table_col2:
            if tab5_prior_forecast_file is not None:
                try:
                    tab5_prior_forecast_df = pd.read_csv(tab5_prior_forecast_file)
                    # Format all columns except Date/Week/Month as number
                    tab5_prior_forecast_df_display = tab5_prior_forecast_df.copy()
                    tab5_prior_forecast_df_display = tab5_prior_forecast_df_display.reset_index(drop=True)
                    for col in tab5_prior_forecast_df_display.columns:
                        if col not in ['Date', 'Week', 'Month']:
                            tab5_prior_forecast_df_display[col] = tab5_prior_forecast_df_display[col].apply(
                                lambda x: f"{x:,.0f}" if pd.notna(x) and isinstance(x, (int, float)) else x
                            )
                    st.markdown("##### Prior Forecast Data Preview")
                    st.dataframe(tab5_prior_forecast_df_display, use_container_width=True, height=480, hide_index=True)
                    st.caption(f"Total Records: {len(tab5_prior_forecast_df):,}")
                except Exception as e:
                    st.error(f"Error reading Prior Forecast CSV: {str(e)}")
                    tab5_prior_forecast_df = None

        # ===== WALK SECTIONS =====
        # Get current date and calculate defaults
        current_date = datetime.now()

        # Calculate quarter begin date
        current_quarter = (current_date.month - 1) // 3 + 1
        quarter_start_month = (current_quarter - 1) * 3 + 1
        quarter_begin_date = datetime(current_date.year, quarter_start_month, 1)

        # Calculate month end date
        if current_date.month == 12:
            month_end_date = datetime(current_date.year, 12, 31)
        else:
            next_month = datetime(current_date.year, current_date.month + 1, 1)
            month_end_date = next_month - pd.Timedelta(days=1)

        # Get Tab4 daily data
        tab4_daily_data = st.session_state.get('fvf_daily_dollar_data', None)

        # Walk 1
        display_walk_section(
            walk_number=1,
            walk_title_default="",
            from_version_default="Prior Forecast",
            to_version_default="Current Forecast",
            start_date_default=quarter_begin_date,
            end_date_default=month_end_date,
            tab4_daily_data=tab4_daily_data,
            tab5_budget_df=tab5_budget_df,
            tab5_prior_forecast_df=tab5_prior_forecast_df
        )

        # Calculate month begin date of next month
        if current_date.month == 12:
            next_month_begin_date = datetime(current_date.year + 1, 1, 1)
        else:
            next_month_begin_date = datetime(current_date.year, current_date.month + 1, 1)

        # Calculate quarter end date
        quarter_end_month = current_quarter * 3
        if quarter_end_month == 12:
            quarter_end_date = datetime(current_date.year, 12, 31)
        else:
            next_month_after_quarter = datetime(current_date.year, quarter_end_month + 1, 1)
            quarter_end_date = next_month_after_quarter - pd.Timedelta(days=1)

        # Walk 2
        display_walk_section(
            walk_number=2,
            walk_title_default="",
            from_version_default="Prior Forecast",
            to_version_default="Current Forecast",
            start_date_default=next_month_begin_date,
            end_date_default=quarter_end_date,
            tab4_daily_data=tab4_daily_data,
            tab5_budget_df=tab5_budget_df,
            tab5_prior_forecast_df=tab5_prior_forecast_df
        )

        # Calculate next quarter begin and end dates
        next_quarter = current_quarter + 1
        if next_quarter > 4:
            # Next quarter is Q1 of next year
            next_quarter_begin_date = datetime(current_date.year + 1, 1, 1)
            next_quarter_end_date = datetime(current_date.year + 1, 3, 31)
        else:
            # Next quarter is in the same year
            next_quarter_start_month = (next_quarter - 1) * 3 + 1
            next_quarter_begin_date = datetime(current_date.year, next_quarter_start_month, 1)

            # Calculate end date of next quarter
            next_quarter_end_month = next_quarter * 3
            if next_quarter_end_month == 12:
                next_quarter_end_date = datetime(current_date.year, 12, 31)
            else:
                month_after_next_quarter = datetime(current_date.year, next_quarter_end_month + 1, 1)
                next_quarter_end_date = month_after_next_quarter - pd.Timedelta(days=1)

        # Walk 3
        display_walk_section(
            walk_number=3,
            walk_title_default="",
            from_version_default="Prior Forecast",
            to_version_default="Current Forecast",
            start_date_default=next_quarter_begin_date,
            end_date_default=next_quarter_end_date,
            tab4_daily_data=tab4_daily_data,
            tab5_budget_df=tab5_budget_df,
            tab5_prior_forecast_df=tab5_prior_forecast_df
        )

    with tab6:
        # ===== SECTION 1: PLUG TABLE =====
        st.markdown("### 🔌 PLUG")

        # Initialize PLUG table in session state if not exists
        if 'plug_table_input' not in st.session_state:
            # Create default PLUG table with 12 months, all values as null
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            st.session_state['plug_table_input'] = pd.DataFrame({
                'Month': months,
                'FVF (%)': [None] * 12,
                'eTRS (%)': [None] * 12,
                'IF ($)': [None] * 12,
                'FF ($)': [None] * 12,
                'Store ($)': [None] * 12
            })

        # Buttons for PLUG table
        plug_col1, plug_col2 = st.columns([1, 5])
        with plug_col1:
            if st.button("🗑️ Clear Cache", key="plug_clear_cache"):
                months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                st.session_state['plug_table_input'] = pd.DataFrame({
                    'Month': months,
                    'FVF (%)': [None] * 12,
                    'eTRS (%)': [None] * 12,
                    'IF ($)': [None] * 12,
                    'FF ($)': [None] * 12,
                    'Store ($)': [None] * 12
                })
                if 'month_to_plug' in st.session_state:
                    del st.session_state['month_to_plug']
                st.rerun()

        with plug_col2:
            if st.button("📥 Read Input", key="plug_read_input"):
                # Store PLUG values from input table
                if 'plug_table_input' in st.session_state:
                    plug_data = st.session_state['plug_table_input'].copy()

                    # Create month to PLUG mapping
                    month_to_plug = {}
                    for idx, row in plug_data.iterrows():
                        month_name = row['Month']
                        month_num = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'].index(month_name) + 1
                        # User enters percentage values (e.g., 5.0 for 5%), TR table uses decimals (0.05 for 5%)
                        # So divide by 100 to convert
                        fvf_plug = row['FVF (%)'] / 100 if pd.notna(row['FVF (%)']) else 0
                        etrs_plug = row['eTRS (%)'] / 100 if pd.notna(row['eTRS (%)']) else 0
                        # Dollar amounts stay as-is
                        if_plug = row['IF ($)'] if pd.notna(row['IF ($)']) else 0
                        ff_plug = row['FF ($)'] if pd.notna(row['FF ($)']) else 0
                        store_plug = row['Store ($)'] if pd.notna(row['Store ($)']) else 0

                        month_to_plug[month_num] = {
                            'FVF': fvf_plug,
                            'eTRS': etrs_plug,
                            'IF ($)': if_plug,
                            'FF ($)': ff_plug,
                            'Store ($)': store_plug
                        }

                    st.session_state['month_to_plug'] = month_to_plug
                    st.success("✅ PLUG values read successfully and applied!")
                    st.rerun()

        # Always show editable PLUG table
        edited_plug_table = st.data_editor(
            st.session_state['plug_table_input'],
            use_container_width=True,
            hide_index=True,
            height=460,
            column_config={
                'Month': st.column_config.TextColumn('Month', disabled=True),
                'FVF (%)': st.column_config.NumberColumn('FVF (%)', format="%.3f%%"),
                'eTRS (%)': st.column_config.NumberColumn('eTRS (%)', format="%.3f%%"),
                'IF ($)': st.column_config.NumberColumn('IF ($)', format="%,.0f"),
                'FF ($)': st.column_config.NumberColumn('FF ($)', format="%,.0f"),
                'Store ($)': st.column_config.NumberColumn('Store ($)', format="%,.0f")
            }
        )
        # Update session state with edited values
        st.session_state['plug_table_input'] = edited_plug_table

        # ===== SECTION 2: DAILY CONSOLIDATION AFTER PLUG =====
        st.markdown("---")
        st.markdown("### 📅 Daily Consolidation after PLUG")
        st.markdown("#### Revenue Dollar Amount")

        # Buttons for Revenue Dollar Amount table
        daily_plug_col1, daily_plug_col2 = st.columns([1, 5])
        with daily_plug_col1:
            if st.button("🗑️ Clear Cache", key="daily_plug_clear_cache"):
                # Reset revenue table to original Tab4 values
                if 'edited_daily_after_plug_data' in st.session_state:
                    del st.session_state['edited_daily_after_plug_data']
                st.success("✅ Revenue table cleared and reset to original values!")
                st.rerun()

        with daily_plug_col2:
            if st.button("📥 Read Input", key="daily_plug_read_input"):
                # Process user GMV edits and store them
                if 'daily_after_plug_editor' in st.session_state:
                    st.session_state['edited_daily_after_plug_data'] = st.session_state['daily_after_plug_editor']
                    st.success("✅ Revenue table values read and applied successfully!")
                    st.rerun()

        # Get data from Tab4 daily section
        if 'fvf_daily_dollar_data' in st.session_state and st.session_state['fvf_daily_dollar_data'] is not None:
            daily_data = st.session_state['fvf_daily_dollar_data'].copy()

            # Define columns to keep
            columns_to_keep = ['Date', 'Week', 'Month', 'A/F', 'GMV', 'Net FVF', 'eTRS', 'FVF excl. eTRS', 'IF', 'FF', 'Store']
            available_columns = [col for col in columns_to_keep if col in daily_data.columns]

            if available_columns:
                daily_after_plug = daily_data[available_columns].copy()

                # Apply PLUG adjustments for A/F=F rows
                if 'month_to_plug' in st.session_state and 'fvf_daily_pct_data' in st.session_state:
                    month_to_plug = st.session_state['month_to_plug']
                    daily_tr_data = st.session_state['fvf_daily_pct_data'].copy()

                    # Calculate monthly sums for proportionate PLUG (IF and FF)
                    monthly_if_sums = {}
                    monthly_ff_sums = {}

                    for month_num in range(1, 13):
                        month_rows_f = daily_after_plug[(daily_after_plug['Month'] == month_num) & (daily_after_plug['A/F'] == 'F')]
                        if 'IF' in daily_after_plug.columns:
                            monthly_if_sums[month_num] = month_rows_f['IF'].sum()
                        if 'FF' in daily_after_plug.columns:
                            monthly_ff_sums[month_num] = month_rows_f['FF'].sum()

                    # Apply adjustments row by row for A/F=F
                    for idx in range(len(daily_after_plug)):
                        if daily_after_plug.iloc[idx].get('A/F') == 'F':
                            month_num = daily_after_plug.iloc[idx].get('Month')

                            if month_num in month_to_plug:
                                # Get GMV for this row
                                gmv = daily_after_plug.iloc[idx].get('GMV', 0)

                                # Get corresponding TR values from TR table (with PLUG already applied)
                                if idx < len(daily_tr_data):
                                    # Net FVF = TR * GMV
                                    if 'Net FVF' in daily_tr_data.columns and pd.notna(daily_tr_data.iloc[idx]['Net FVF']):
                                        tr_net_fvf = daily_tr_data.iloc[idx]['Net FVF']
                                        # Apply PLUG to TR first
                                        tr_net_fvf += month_to_plug[month_num]['FVF']
                                        daily_after_plug.iloc[idx, daily_after_plug.columns.get_loc('Net FVF')] = tr_net_fvf * gmv

                                    # eTRS = TR * GMV
                                    if 'eTRS' in daily_tr_data.columns and pd.notna(daily_tr_data.iloc[idx]['eTRS']):
                                        tr_etrs = daily_tr_data.iloc[idx]['eTRS']
                                        # Apply PLUG to TR first
                                        tr_etrs += month_to_plug[month_num]['eTRS']
                                        daily_after_plug.iloc[idx, daily_after_plug.columns.get_loc('eTRS')] = tr_etrs * gmv

                                # IF: Tab4 + proportionate PLUG
                                if 'IF' in daily_after_plug.columns and 'IF ($)' in month_to_plug.get(month_num, {}):
                                    original_if = daily_data.iloc[idx]['IF'] if idx < len(daily_data) else 0
                                    plug_if_monthly = month_to_plug[month_num].get('IF ($)', 0)

                                    if monthly_if_sums.get(month_num, 0) > 0:
                                        # Proportionate PLUG
                                        plug_if_daily = plug_if_monthly * (original_if / monthly_if_sums[month_num])
                                    else:
                                        plug_if_daily = 0

                                    daily_after_plug.iloc[idx, daily_after_plug.columns.get_loc('IF')] = original_if + plug_if_daily

                                # FF: Tab4 + proportionate PLUG
                                if 'FF' in daily_after_plug.columns and 'FF ($)' in month_to_plug.get(month_num, {}):
                                    original_ff = daily_data.iloc[idx]['FF'] if idx < len(daily_data) else 0
                                    plug_ff_monthly = month_to_plug[month_num].get('FF ($)', 0)

                                    if monthly_ff_sums.get(month_num, 0) > 0:
                                        # Proportionate PLUG
                                        plug_ff_daily = plug_ff_monthly * (original_ff / monthly_ff_sums[month_num])
                                    else:
                                        plug_ff_daily = 0

                                    daily_after_plug.iloc[idx, daily_after_plug.columns.get_loc('FF')] = original_ff + plug_ff_daily

                                # Store: Direct PLUG on 1st of month only
                                if 'Store' in daily_after_plug.columns and 'Store ($)' in month_to_plug.get(month_num, {}):
                                    date_val = daily_after_plug.iloc[idx].get('Date')
                                    if pd.notna(date_val) and date_val.day == 1:
                                        original_store = daily_data.iloc[idx]['Store'] if idx < len(daily_data) else 0
                                        plug_store = month_to_plug[month_num].get('Store ($)', 0)
                                        daily_after_plug.iloc[idx, daily_after_plug.columns.get_loc('Store')] = original_store + plug_store

                # Create display version
                daily_display = daily_after_plug.copy()

                # Format Date to short format
                if 'Date' in daily_display.columns:
                    daily_display['Date'] = daily_display['Date'].apply(
                        lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else ""
                    )

                # Format numeric columns with thousand separators for all rows (including A/F=A)
                numeric_cols = ['GMV', 'Net FVF', 'eTRS', 'FVF excl. eTRS', 'IF', 'FF', 'Store']
                for col in numeric_cols:
                    if col in daily_display.columns:
                        daily_display[col] = daily_display[col].apply(
                            lambda x: f"{x:,.0f}" if pd.notna(x) else ""
                        )

                # Apply styling - light grey background for A/F=A rows, EXCLUDING GMV column
                def highlight_actual_rows_except_gmv(row):
                    colors = []
                    for col in daily_display.columns:
                        if row.get('A/F') == 'A' and col != 'GMV':
                            colors.append('background-color: #f0f0f0')
                        else:
                            colors.append('')
                    return colors

                styled_daily_display = daily_display.style.apply(highlight_actual_rows_except_gmv, axis=1)

                # Configure column settings
                daily_column_config = {
                    "Date": st.column_config.TextColumn("Date", disabled=True),
                    "Week": st.column_config.TextColumn("Week", disabled=True),
                    "Month": st.column_config.TextColumn("Month", disabled=True),
                    "A/F": st.column_config.TextColumn("A/F", disabled=True, help="A = Actual, F = Forecast"),
                    "GMV": st.column_config.TextColumn("GMV", help="Editable for Forecast rows only. Actual rows (A/F=A) are protected."),
                }
                for col in ['Net FVF', 'eTRS', 'FVF excl. eTRS', 'IF', 'FF', 'Store']:
                    if col in daily_display.columns:
                        daily_column_config[col] = st.column_config.TextColumn(col, disabled=True)

                # Display editable table with styling
                edited_daily_after_plug = st.data_editor(
                    styled_daily_display,
                    use_container_width=True,
                    height=600,
                    hide_index=True,
                    column_config=daily_column_config,
                    key="daily_after_plug_editor"
                )

                st.caption(f"Total Records: {len(daily_display):,}")
            else:
                st.warning("⚠️ Required columns not found in daily data.")
        else:
            st.info("ℹ️ Daily data not available. Please generate forecast in Tab4 first.")

        # ===== SECTION 3: TAKE RATE % =====
        st.markdown("#### Take Rate %")

        # Get GMV from Revenue Dollar Amount table above (edited_daily_after_plug)
        # Get TR percentages from Tab4 daily section
        if 'fvf_daily_pct_data' in st.session_state and st.session_state['fvf_daily_pct_data'] is not None:
            daily_tr_data = st.session_state['fvf_daily_pct_data'].copy()

            # Define columns to keep
            tr_columns_to_keep = ['Date', 'Week', 'Month', 'A/F', 'GMV', 'Net FVF', 'eTRS', 'FVF excl. eTRS', 'IF', 'FF', 'Store']

            # Filter columns (only keep columns that exist in the data)
            available_tr_columns = [col for col in tr_columns_to_keep if col in daily_tr_data.columns]

            if available_tr_columns:
                daily_tr_table = daily_tr_data[available_tr_columns].copy()

                # Extract values from Revenue Dollar Amount table to calculate TR
                if 'edited_daily_after_plug' in locals() and edited_daily_after_plug is not None:
                    # Parse all revenue values from the displayed table
                    revenue_values = {}
                    for col in ['GMV', 'Net FVF', 'eTRS', 'IF', 'FF', 'Store']:
                        if col in edited_daily_after_plug.columns:
                            revenue_values[col] = []
                            for idx in range(len(edited_daily_after_plug)):
                                value_str = edited_daily_after_plug.iloc[idx][col]
                                if value_str and str(value_str).strip():
                                    try:
                                        # Remove commas and parse as float
                                        value = float(str(value_str).replace(',', '').strip())
                                        revenue_values[col].append(value)
                                    except ValueError:
                                        revenue_values[col].append(0)
                                else:
                                    revenue_values[col].append(0)

                    # Update TR table with GMV from revenue table
                    if 'GMV' in revenue_values:
                        daily_tr_table['GMV'] = revenue_values['GMV']

                    # Calculate TR for each row
                    for idx in range(len(daily_tr_table)):
                        gmv = revenue_values['GMV'][idx] if idx < len(revenue_values['GMV']) else 0
                        af_status = daily_tr_table.iloc[idx].get('A/F', 'F')

                        if gmv > 0:
                            if af_status == 'F':  # Forecast rows: apply PLUG logic
                                # Net FVF: Tab4 + PLUG
                                if 'Net FVF' in daily_tr_table.columns and 'month_to_plug' in st.session_state:
                                    month_num = daily_tr_table.iloc[idx].get('Month')
                                    month_to_plug = st.session_state['month_to_plug']

                                    if month_num in month_to_plug:
                                        original_fvf = daily_tr_data.iloc[idx]['Net FVF'] if idx < len(daily_tr_data) else 0
                                        adjusted_fvf = original_fvf + month_to_plug[month_num]['FVF']
                                        daily_tr_table.iloc[idx, daily_tr_table.columns.get_loc('Net FVF')] = adjusted_fvf

                                        # eTRS: Tab4 + PLUG
                                        if 'eTRS' in daily_tr_table.columns:
                                            original_etrs = daily_tr_data.iloc[idx]['eTRS'] if idx < len(daily_tr_data) else 0
                                            adjusted_etrs = original_etrs + month_to_plug[month_num]['eTRS']
                                            daily_tr_table.iloc[idx, daily_tr_table.columns.get_loc('eTRS')] = adjusted_etrs

                                # IF/FF/Store: Revenue table / GMV
                                if 'IF' in daily_tr_table.columns and 'IF' in revenue_values:
                                    if_revenue = revenue_values['IF'][idx]
                                    daily_tr_table.iloc[idx, daily_tr_table.columns.get_loc('IF')] = if_revenue / gmv

                                if 'FF' in daily_tr_table.columns and 'FF' in revenue_values:
                                    ff_revenue = revenue_values['FF'][idx]
                                    daily_tr_table.iloc[idx, daily_tr_table.columns.get_loc('FF')] = ff_revenue / gmv

                                if 'Store' in daily_tr_table.columns and 'Store' in revenue_values:
                                    store_revenue = revenue_values['Store'][idx]
                                    daily_tr_table.iloc[idx, daily_tr_table.columns.get_loc('Store')] = store_revenue / gmv
                            # For Actual rows (A/F=A), keep original TR values from Tab4

                # Create display version
                tr_display = daily_tr_table.copy()

                # Format Date to short format
                if 'Date' in tr_display.columns:
                    tr_display['Date'] = tr_display['Date'].apply(
                        lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else ""
                    )

                # Format GMV with thousand separators and no decimals
                if 'GMV' in tr_display.columns:
                    tr_display['GMV'] = tr_display['GMV'].apply(
                        lambda x: f"{x:,.0f}" if pd.notna(x) else ""
                    )

                # Format percentage columns (Net FVF to Store) with 2 decimals
                # Note: values are in decimal format (0.1234 = 12.34%), so multiply by 100
                pct_cols = ['Net FVF', 'eTRS', 'FVF excl. eTRS', 'IF', 'FF', 'Store']
                for col in pct_cols:
                    if col in tr_display.columns:
                        tr_display[col] = tr_display[col].apply(
                            lambda x: f"{x*100:.2f}%" if pd.notna(x) else ""
                        )

                # Apply styling - light grey background for A/F=A rows (all columns)
                def highlight_actual_rows(row):
                    if row.get('A/F') == 'A':
                        return ['background-color: #f0f0f0'] * len(row)
                    else:
                        return [''] * len(row)

                styled_tr_display = tr_display.style.apply(highlight_actual_rows, axis=1)

                # Display TR table
                st.dataframe(styled_tr_display, use_container_width=True, height=600, hide_index=True)
                st.caption(f"Total Records: {len(tr_display):,}")
            else:
                st.warning("⚠️ Required columns not found in TR data.")
        else:
            st.info("ℹ️ TR data not available. Please generate forecast in Tab4 first.")

        # ===== SECTION 4: MONTHLY CONSOLIDATION AFTER PLUG =====
        st.markdown("---")
        st.markdown("### 📊 Monthly Consolidation after PLUG")

        # Calculate monthly revenue table by aggregating daily revenue table
        if 'daily_after_plug' in locals() and daily_after_plug is not None:
            # Aggregate by Month
            monthly_revenue = daily_after_plug.groupby('Month').agg({
                'GMV': 'sum',
                'Net FVF': 'sum',
                'eTRS': 'sum',
                'IF': 'sum',
                'FF': 'sum',
                'Store': 'sum'
            }).reset_index()

            # Add month names
            month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            monthly_revenue['Month Name'] = monthly_revenue['Month'].apply(lambda x: month_names[x-1] if 1 <= x <= 12 else '')

            # Reorder columns: Month Name first
            cols = ['Month Name', 'GMV', 'Net FVF', 'eTRS', 'IF', 'FF', 'Store']
            monthly_revenue = monthly_revenue[cols]

            # Display Monthly Revenue Table
            st.markdown("#### Revenue Dollar Amount")

            # Format for display
            monthly_revenue_display = monthly_revenue.copy()
            for col in ['GMV', 'Net FVF', 'eTRS', 'IF', 'FF', 'Store']:
                if col in monthly_revenue_display.columns:
                    monthly_revenue_display[col] = monthly_revenue_display[col].apply(
                        lambda x: f"{x:,.0f}" if pd.notna(x) else ""
                    )

            st.dataframe(monthly_revenue_display, use_container_width=True, hide_index=True, height=470)

            # Calculate Monthly TR Table
            st.markdown("#### Take Rate %")

            monthly_tr = monthly_revenue.copy()

            # Calculate TR for each column: revenue / GMV
            for col in ['Net FVF', 'eTRS', 'IF', 'FF', 'Store']:
                if col in monthly_tr.columns:
                    monthly_tr[col] = monthly_tr.apply(
                        lambda row: (row[col] / row['GMV']) if row['GMV'] > 0 else 0,
                        axis=1
                    )

            # Format for display
            monthly_tr_display = monthly_tr.copy()

            # Format GMV with thousand separators
            monthly_tr_display['GMV'] = monthly_tr_display['GMV'].apply(
                lambda x: f"{x:,.0f}" if pd.notna(x) else ""
            )

            # Format TR columns as percentages (multiply by 100)
            for col in ['Net FVF', 'eTRS', 'IF', 'FF', 'Store']:
                if col in monthly_tr_display.columns:
                    monthly_tr_display[col] = monthly_tr_display[col].apply(
                        lambda x: f"{x*100:.2f}%" if pd.notna(x) else ""
                    )

            st.dataframe(monthly_tr_display, use_container_width=True, hide_index=True, height=470)

        else:
            st.info("ℹ️ Monthly consolidation data not available. Please generate daily data first.")


if __name__ == "__main__":
    main()
